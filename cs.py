import pandas as pd
from openpyxl import load_workbook
import xlsxwriter

wb = load_workbook('./data.xlsx')
sheetArr =wb.sheetnames

workbook = xlsxwriter.Workbook('processed.xlsx')
worksheet = workbook.add_worksheet()

banDevi = wb[sheetArr[4]]
print('Working Sheet', wb[sheetArr[4]])

class Student:
    def __init__(self, name, age, grade, gender, litnumNepaliB, litnumNepaliM,  litnumNepaliP, litnumMathB, litnumMathM, litnumMathP):
        self.name = name
        self.age = age
        self.grade = grade
        self.gender = gender
        self.litnumNepaliB = litnumNepaliB
        self.litnumNepaliM = litnumNepaliM
        self.litnumNepaliP = litnumNepaliP
        self.litnumMathB = litnumMathB
        self.litnumMathM = litnumMathM
        self.litnumMathP = litnumMathP

def calculateLitnum(term, startChar):
    for j in range(5):
        #for litnum Nepali
        if banDevi[startChar+str(term)].value is not None:
            litValue = j+1
            return litValue
        startChar = chr(ord(startChar)+1)

nameCount, ageCount, baseCount, midCount, postCount = 2, 5, 6, 7, 8
students = []
i = 0
while banDevi['C'+str(nameCount)].value is not None:
    if banDevi['D'+str(ageCount)].value is None:
        gender = 'M'
    else:
        gender = 'F'
    litNepaliB = calculateLitnum(baseCount, 'F')
    litNepaliM = calculateLitnum(midCount, 'F')
    litNepaliP = calculateLitnum(postCount, 'F')
    litMathB = calculateLitnum(baseCount, 'K')
    litMathM = calculateLitnum(midCount, 'K')
    litMathP = calculateLitnum(postCount, 'K')
    wpmCountB = banDevi['P'+str(baseCount)].value if banDevi['P'+str(baseCount)].value is not None else 0
    wpmCountP = banDevi['P'+str(midCount)].value if banDevi['P'+str(midCount)].value is not None else 0
    wpmCountM = banDevi['P'+str(postCount)].value if banDevi['P'+str(postCount)].value is not None else 0
    student = Student(banDevi['C'+str(nameCount)].value, banDevi['C'+str(ageCount)].value, banDevi['B'+str(ageCount)].value, gender, litNepaliB, litNepaliM, litNepaliP, litMathB, litMathM, litMathP)
    students.append(student)
    nameCount += 8
    ageCount += 8
    baseCount += 8
    midCount += 8
    postCount += 8
    print(students[i].name, students[i].age, students[i].grade, students[i].gender, students[i].litnumNepaliB, students[i].litnumNepaliM, students[i].litnumNepaliP, students[i].litnumMathB, students[i].litnumMathM, students[i].litnumMathP)
    i += 1

row, col = 1, 0
for obj in students:
    worksheet.write_string(row, col, obj.name)
    worksheet.write_string(row, col+ 1, str(obj.litnumNepaliB))
    worksheet.write_string(row, col+ 2, str(obj.litnumMathB))
    worksheet.write_string(row, col+ 3, str(obj.litnumNepaliM))
    worksheet.write_string(row, col+ 4, str(obj.litnumMathM))
    worksheet.write_string(row, col+ 5, str(obj.litnumNepaliP))
    worksheet.write_string(row, col+ 6, str(obj.litnumMathP))
    worksheet.write_string(row, col+ 7, str(obj.grade))
    worksheet.write_string(row, col+ 8, str(obj.age))
    row += 1

workbook.close()
